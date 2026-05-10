"""Excel report – jeden list s přehledem, W3C, strukturou, robots.txt, uživatelskou sekcí."""
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from issues import Issue, IssueType
from robots_check import CRITICAL_PREFIX as ROBOTS_CRITICAL_PREFIX
from stats  import compute_stats
from ui     import is_local_url

# ── Paleta ───────────────────────────────────────────────────────────────────
G_BG="C6EFCE"; G_FT="1E6823"
R_BG="FFC7CE"; R_FT="9C0006"
O_BG="FFEB9C"; O_FT="7D5A00"
B_BG="DDEEFF"; B_FT="1F4E79"
GR_BG="F5F5F5"; GR_FT="555555"
HDR="1F3864"; SUB="2E75B6"; SUB2="1A5276"

SCORE_COLORS = {
    "good": ("C6EFCE", "1E6823"),
    "warn": ("FFEB9C", "7D5A00"),
    "bad":  ("FFC7CE", "9C0006"),
}


# ── Stylovací helpers ─────────────────────────────────────────────────────────
def _hf(sz=10, color="FFFFFF"):
    return Font(name="Arial", bold=True, color=color, size=sz)

def _bf(bold=False, color="222222", sz=10):
    return Font(name="Arial", bold=bold, color=color, size=sz)

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _al(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def _brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _dc(ws, r, c, v, bold=False, bg=None, ft=None, align="left", sz=10, link=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _bf(bold=bold, color=ft or "222222", sz=sz)
    cell.alignment = _al(align)
    cell.border = _brd()
    if bg:   cell.fill = _fill(bg)
    if link:
        cell.hyperlink = link
        cell.font = Font(name="Arial", color="1155CC", underline="single", sz=sz)
    return cell

def _badge(ws, r, c, txt, bg, ft):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font = _bf(bold=True, color=ft, sz=10)
    cell.fill = _fill(bg); cell.alignment = _al("center"); cell.border = _brd()

def _title_row(ws, row, text, bg, end_col="G", sz=11):
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = _hf(sz); c.fill = _fill(bg); c.alignment = _al("center")
    ws.row_dimensions[row].height = 24

def _hdr_row(ws, cols_def, row, bg=None):
    for i, (title, _) in enumerate(cols_def, 1):
        c = ws.cell(row=row, column=i, value=title)
        c.font = _hf(10); c.fill = _fill(bg or HDR)
        c.alignment = _al("center"); c.border = _brd()
    ws.row_dimensions[row].height = 22

def _spacer(ws, row, height=10):
    ws.row_dimensions[row].height = height
    return row + 1

def _as_https(url: str) -> str:
    """
    Normalizuje URL na https:// (přepíše http:// → https://).
    Lokální URL (localhost, IP, *.local…) se nemění — tam je http:// správné
    a https varianta typicky neexistuje.
    """
    if not isinstance(url, str):
        return url
    if not url.startswith("http://"):
        return url
    if is_local_url(url):
        return url
    return "https://" + url[7:]

def _w3c_link(url: str) -> str:
    return "https://validator.w3.org/nu/?doc=" + quote(_as_https(url), safe="")

def _score_palette(score: int) -> tuple[str, str]:
    if score >= 80: return SCORE_COLORS["good"]
    if score >= 60: return SCORE_COLORS["warn"]
    return SCORE_COLORS["bad"]


# ── Sekce: souhrn + meta ──────────────────────────────────────────────────────

def _write_header(ws, start_url: str, source_label: str) -> int:
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "WEB VALIDATOR – REPORT"
    t.font  = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    t.fill  = _fill(HDR); t.alignment = _al("center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:G2")
    m = ws["A2"]
    src_info = f"  |  Zdroj: {source_label}" if source_label else ""
    m.value = (f"Web: {_as_https(start_url)}     |     Datum: "
               f"{datetime.now().strftime('%d.%m.%Y  %H:%M')}{src_info}")
    m.font = Font(name="Arial", size=10, color="555555"); m.alignment = _al("center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10
    return 4


def _write_summary(ws, row: int, score: int, stats, http_converted: bool = False) -> int:
    _title_row(ws, row, "SOUHRN", SUB); row += 1

    if score >= 0:
        sc_bg, sc_ft = _score_palette(score)
        sc_label = "Výborný" if score >= 80 else "Průměrný" if score >= 60 else "Špatný"
        ws.merge_cells(f"A{row}:E{row}")
        lc = ws.cell(row=row, column=1, value="Web Quality Score")
        lc.font = Font(name="Arial", bold=True, size=13)
        lc.alignment = _al(); lc.border = _brd()
        ws.merge_cells(f"F{row}:G{row}")
        vc = ws.cell(row=row, column=6, value=f"{score}/100  –  {sc_label}")
        vc.font = Font(name="Arial", bold=True, size=14, color=sc_ft)
        vc.fill = _fill(sc_bg); vc.alignment = _al("center"); vc.border = _brd()
        ws.row_dimensions[row].height = 32; row += 1

    cards = [
        ("Zkontrolováno stránek", stats.total,     B_BG, B_FT),
        ("W3C – Bez problémů",    stats.w3c_ok,    G_BG, G_FT),
        ("W3C – Varování",        stats.w3c_warn,
            O_BG if stats.w3c_warn else G_BG, O_FT if stats.w3c_warn else G_FT),
        ("W3C – Chyby",           stats.w3c_err,
            R_BG if stats.w3c_err else G_BG, R_FT if stats.w3c_err else G_FT),
        ("Struktura – OK",        stats.struct_ok, G_BG, G_FT),
        ("Struktura – Problémy",  stats.struct_bad,
            O_BG if stats.struct_bad else G_BG, O_FT if stats.struct_bad else G_FT),
    ]
    if stats.w3c_failed:
        cards.append(("Nepodařilo načíst stránek", stats.w3c_failed, R_BG, R_FT))

    for label, val, bg, ft in cards:
        ws.merge_cells(f"A{row}:E{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _bf(bold=True, sz=11); lc.alignment = _al(); lc.border = _brd()
        ws.merge_cells(f"F{row}:G{row}")
        vc = ws.cell(row=row, column=6, value=val)
        vc.font = Font(name="Arial", bold=True, size=13, color=ft)
        vc.fill = _fill(bg); vc.alignment = _al("center"); vc.border = _brd()
        ws.row_dimensions[row].height = 22; row += 1

    # Upozornění na http → https převod (zobrazí se jen když k tomu skutečně došlo
    # a převod se týkal veřejných URL — pro lokální host je http:// správné).
    if http_converted:
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1,
                       value="ℹ Sitemap obsahovala URL s http:// – v reportu jsou zobrazeny jako https://")
        cell.font = _bf(bold=False, color=B_FT, sz=10)
        cell.fill = _fill(B_BG)
        cell.alignment = _al("center")
        cell.border = _brd()
        ws.row_dimensions[row].height = 20
        row += 1

    return row


def _write_homepage_meta(ws, row: int, results: list) -> int:
    row = _spacer(ws, row)
    _title_row(ws, row, "META – HOMEPAGE", SUB); row += 1

    # Najdi první výsledek s neprázdnou homepage_meta (může být kdekoli v results,
    # ne nutně na indexu 0 — sitemap může mít homepage uprostřed nebo úplně chybět).
    # main.py nastaví homepage_meta jen na URL která odpovídá startovní URL auditu.
    homepage_meta = next(
        (r["homepage_meta"] for r in results if r.get("homepage_meta")),
        []
    )

    if not homepage_meta:
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value="Žádná data").font = _bf(color="888888")
        ws.row_dimensions[row].height = 18
        return row + 1

    for line in homepage_meta:
        line = line.strip()
        if not line: continue
        if "v pořádku" in line:
            bg2, ft2, badge = G_BG, G_FT, "OK"
        elif any(k in line for k in ["příliš", "Chybí"]):
            bg2, ft2, badge = R_BG, R_FT, "PROBLÉM"
        else:
            bg2, ft2, badge = GR_BG, GR_FT, "INFO"
        if ":" in line and not line.startswith(" "):
            key, val2 = line.split(":", 1)[0].strip(), line.split(":", 1)[1].strip()
        else:
            key, val2 = line, ""
        ws.merge_cells(f"A{row}:B{row}"); _dc(ws, row, 1, key, bold=True)
        ws.merge_cells(f"C{row}:E{row}"); _dc(ws, row, 3, val2)
        ws.merge_cells(f"F{row}:G{row}"); _badge(ws, row, 6, badge, bg2, ft2)
        ws.row_dimensions[row].height = 20; row += 1

    return row


def _format_w3c_messages(warnings: list, errors: list, max_items: int = 30) -> str:
    """
    Sestaví textový výpis W3C zpráv pro Excel buňku.

    Vstup: list dictů ve formátu {"message": str, "line": int|None}
           (formát použitý ve validator_w3c._classify).
    Výstup: víceřádkový text s předponou typu a čísla řádku, např.:
        [CHYBA @ 23] Element "img" missing required attribute "alt".
        [VAROVÁNÍ @ 41] Section lacks heading.

    Chyby jsou nahoře (závažnější), pak varování. Při velkém počtu zpráv
    se zobrazí max_items a zbytek se shrne řádkem '… a další N'.
    """
    lines: list[str] = []

    def _format_one(prefix: str, item: dict) -> str:
        # `item` je dict {"message": "...", "line": int|None}
        # Defenzivně: kdyby přišel string (starší formát), zachovej ho.
        if isinstance(item, str):
            return f"[{prefix}] {item}"
        msg = (item.get("message") or "").strip()
        ln  = item.get("line")
        loc = f" @ ř.{ln}" if ln else ""
        return f"[{prefix}{loc}] {msg}"

    for e in errors:
        lines.append(_format_one("CHYBA", e))
    for w in warnings:
        lines.append(_format_one("VAROVÁNÍ", w))

    if len(lines) > max_items:
        rest = len(lines) - max_items
        lines = lines[:max_items] + [f"… a dalších {rest} zpráv (oříznuto)"]

    return "\n".join(lines) if lines else ""


def _write_w3c_section(ws, row: int, results: list, is_local_audit: bool = False) -> int:
    """
    Souhrn W3C validace.

    Pro VEŘEJNÉ weby zobrazuje pro každou problematickou stránku odkaz
    na https://validator.w3.org/nu/?doc=URL — uživatel klikne a uvidí detail.

    Pro LOKÁLNÍ audit (localhost, 127.0.0.1, *.local…) tento odkaz nefunguje,
    protože W3C server lokální URL nedohledá. Místo odkazu zobrazíme:
      - přímou URL stránky (klikatelnou — uživatel si ji otevře v prohlížeči)
      - sloupec navíc 'Výpis chyby z w3c' s textem všech chyb a varování
        získaných z lokálního vnu.jar
    """
    row = _spacer(ws, row)
    _title_row(ws, row, "W3C VALIDACE – STRÁNKY S PROBLÉMY", SUB); row += 1

    w3c_issues = [r for r in results
                  if r["w3c_category"] not in ("ok", "validator_error")]
    if not w3c_issues:
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value="✓ Žádné W3C problémy nalezeny")
        ws.cell(row=row, column=1).font = _bf(color=G_FT, bold=True)
        ws.row_dimensions[row].height = 20
        return row + 1

    # ── Hlavička tabulky ──────────────────────────────────────────────────────
    # Layout sdílíme s ostatními sekcemi, šířky sloupců jsou globální:
    # A=80, B=12, C=8, D=8, E=30, F=14, G=40
    # Pro veřejný audit:  [W3C URL | Status | Varování | Chyby]   (E:G prázdné)
    # Pro lokální audit:  [Stránka | Status | Varování | Chyby | Výpis chyby z w3c]
    #                                                            └── E:G merged
    if is_local_audit:
        # Hlavička s 5 logickými sloupci (poslední je merged E:G)
        for ci, label in enumerate(
            ["Stránka", "Status", "Varování", "Chyby", "Výpis chyby z w3c"], 1
        ):
            target_col = ci if ci <= 4 else 5
            c = ws.cell(row=row, column=target_col, value=label)
            c.font = _hf(10); c.fill = _fill(SUB2)
            c.alignment = _al("center"); c.border = _brd()
        # E:G merged kvůli šířce výpisu
        ws.merge_cells(f"E{row}:G{row}")
        for col in range(6, 8):
            ws.cell(row=row, column=col).fill = _fill(SUB2)
            ws.cell(row=row, column=col).border = _brd()
        ws.row_dimensions[row].height = 22
    else:
        _hdr_row(ws, [("W3C Validator URL", 80), ("Status", 13),
                      ("Varování", 9), ("Chyby", 8)], row=row, bg=SUB2)
    row += 1

    for r in w3c_issues:
        cat = r["w3c_category"]
        w_txt = "VAROVÁNÍ" if cat == "warning" else "CHYBA" if cat == "error" else "VAR+CHYBA"
        w_bg  = O_BG if cat == "warning" else R_BG
        w_ft  = O_FT if cat == "warning" else R_FT

        if is_local_audit:
            # Přímý odkaz na lokální stránku — uživatel si ji otevře v prohlížeči
            page_url = r["url"]
            _dc(ws, row, 1, page_url, link=page_url)
        else:
            # Veřejný web — odkaz na W3C Validator
            vurl = _w3c_link(r["url"])
            _dc(ws, row, 1, vurl, link=vurl)

        _badge(ws, row, 2, w_txt, w_bg, w_ft)
        _dc(ws, row, 3, len(r["w3c_warnings"]) or "",
            bg=O_BG if r["w3c_warnings"] else None, align="center")
        _dc(ws, row, 4, len(r["w3c_errors"]) or "",
            bg=R_BG if r["w3c_errors"] else None, align="center")

        if is_local_audit:
            # Sloupec navíc — výpis konkrétních chyb a varování z vnu.jar.
            # E:G merged kvůli prostoru (cca 84 znaků šířky).
            messages_text = _format_w3c_messages(
                r.get("w3c_warnings") or [],
                r.get("w3c_errors") or [],
            )
            ws.merge_cells(f"E{row}:G{row}")
            _dc(ws, row, 5, messages_text or "(bez detailů)", align="left", sz=9)
            for col in range(6, 8):
                ws.cell(row=row, column=col).border = _brd()

            # Výška řádku úměrná počtu zpráv (mezi 24 a 200)
            n_lines = len(messages_text.splitlines()) if messages_text else 1
            ws.row_dimensions[row].height = min(200, max(24, 14 * n_lines + 4))
        else:
            ws.row_dimensions[row].height = 18

        row += 1

    return row


def _write_structure_section(ws, row: int, results: list) -> int:
    """
    Souhrn strukturálních problémů.
    Pracuje přímo s Issue objekty — žádný text-parsing!
    """
    row = _spacer(ws, row)
    _title_row(ws, row, "HTML STRUKTURA – SOUHRN PROBLÉMŮ", SUB); row += 1

    # Seskupení podle labelu (Issue.label zahrnuje i konkrétní tag pro EMPTY_TAG)
    grouped: dict[str, list[str]] = defaultdict(list)

    for r in results:
        issues = r.get("structure_issues", [])
        for issue in issues:
            # Ignoruj non-Issue objekty (pro jistotu backward compat)
            if not isinstance(issue, Issue):
                continue
            label = issue.label
            url_https = _as_https(r["url"])
            if url_https not in grouped[label]:
                grouped[label].append(url_https)

    if not grouped:
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value="✓ Žádné strukturální problémy nalezeny")
        ws.cell(row=row, column=1).font = _bf(color=G_FT, bold=True)
        ws.row_dimensions[row].height = 20
        return row + 1

    _hdr_row(ws, [("Typ problému", 35), ("Počet URL", 10),
                  ("Postižené stránky", 110)], row=row, bg=SUB2)
    ws.merge_cells(f"C{row}:G{row}")
    for col in range(4, 8):
        ws.cell(row=row, column=col).fill   = _fill(SUB2)
        ws.cell(row=row, column=col).border = _brd()
    row += 1

    for issue_type, urls in sorted(grouped.items(), key=lambda x: -len(x[1])):
        n   = len(urls)
        bg3 = O_BG if n <= 3 else R_BG
        ft3 = O_FT if n <= 3 else R_FT
        _dc(ws, row, 1, issue_type, bold=True)
        _dc(ws, row, 2, n, bg=bg3, ft=ft3, align="center", bold=True)
        ws.merge_cells(f"C{row}:G{row}")
        _dc(ws, row, 3, "\n".join(urls))
        ws.row_dimensions[row].height = max(18, 15 * min(n, 8))
        row += 1

    return row


def _write_failed_pages(ws, row: int, results: list) -> int:
    failed_pages = [r for r in results if r["w3c_category"] == "validator_error"]
    if not failed_pages:
        return row

    row = _spacer(ws, row)
    _title_row(ws, row, "NEDOSTUPNÉ STRÁNKY – NEPODAŘILO SE NAČÍST", SUB); row += 1

    for ci, label in enumerate(["URL stránky", "", "", "", "Chybová hláška", "", ""], 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = _hf(10); c.fill = _fill(SUB2)
        c.alignment = _al("center"); c.border = _brd()
    ws.merge_cells(f"A{row}:D{row}")
    ws.merge_cells(f"E{row}:G{row}")
    ws.row_dimensions[row].height = 22; row += 1

    for r in failed_pages:
        ws.merge_cells(f"A{row}:D{row}")
        _dc(ws, row, 1, _as_https(r["url"]), bg=R_BG, ft=R_FT)
        ws.merge_cells(f"E{row}:G{row}")
        err_msg = str(r.get("w3c_error_msg") or "")[:200]
        _dc(ws, row, 5, err_msg, bg=R_BG, ft=R_FT)
        ws.row_dimensions[row].height = 20; row += 1

    return row


def _write_robots_section(ws, row: int, robots_issues: list, robots_skipped: bool) -> int:
    row = _spacer(ws, row)
    _title_row(ws, row, "ROBOTS.TXT – BLOKOVÁNÍ INDEXACE / JS / CSS (GOOGLEBOT)", SUB); row += 1

    if robots_skipped:
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1,
                value="ℹ Kontrola přeskočena – interní / dev / lokální prostředí")
        ws.cell(row=row, column=1).font = _bf(color=GR_FT)
        ws.cell(row=row, column=1).fill = _fill(GR_BG)
        ws.cell(row=row, column=1).border = _brd()
        ws.row_dimensions[row].height = 20
        return row + 1
    elif not robots_issues:
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1,
                value="✓ Robots.txt neblokuje indexaci, JS ani CSS – Googlebot může renderovat stránku")
        ws.cell(row=row, column=1).font = _bf(color=G_FT, bold=True)
        ws.row_dimensions[row].height = 20
        return row + 1

    _hdr_row(ws, [("Zjištěný problém v robots.txt", 110)], row=row, bg=SUB2)
    ws.merge_cells(f"B{row}:G{row}")
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = _fill(SUB2)
        ws.cell(row=row, column=col).border = _brd()
    row += 1
    for issue in robots_issues:
        ws.merge_cells(f"A{row}:G{row}")
        is_critical = issue.startswith(ROBOTS_CRITICAL_PREFIX)
        if is_critical:
            # Kritická chyba — Disallow: / blokuje celý web.
            # Větší písmo, tučné, výraznější červená, vyšší řádek.
            clean_msg = "⛔ KRITICKÉ: " + issue[len(ROBOTS_CRITICAL_PREFIX):]
            cell = ws.cell(row=row, column=1, value=clean_msg)
            cell.font = Font(name="Arial", bold=True, color=R_FT, size=11)
            cell.fill = _fill(R_BG)
            cell.alignment = _al("left")
            cell.border = _brd()
            ws.row_dimensions[row].height = 28
        else:
            _dc(ws, row, 1, issue, bg=R_BG, ft=R_FT)
            ws.row_dimensions[row].height = 20
        row += 1

    return row


def _write_user_pages(ws, row: int, user_pages: list, is_local_audit: bool = False) -> int:
    row = _spacer(ws, row)
    _title_row(ws, row, "UŽIVATELSKÁ SEKCE – DETEKCE", SUB); row += 1

    if not user_pages:
        ws.merge_cells(f"A{row}:G{row}")
        if is_local_audit:
            msg = "ℹ Kontrola přeskočena – lokální / privátní host"
        else:
            msg = "Kontrola neproběhla nebo nedostupná"
        ws.cell(row=row, column=1, value=msg)
        ws.cell(row=row, column=1).font = _bf(color=GR_FT)
        ws.cell(row=row, column=1).fill = _fill(GR_BG) if is_local_audit else PatternFill()
        ws.row_dimensions[row].height = 18
        return row + 1

    for ci, label in enumerate(["Cesta", "URL", "", "", "HTTP Status", "", "Existence"], 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = _hf(10); c.fill = _fill(SUB2)
        c.alignment = _al("center"); c.border = _brd()
    ws.merge_cells(f"B{row}:D{row}")
    ws.merge_cells(f"E{row}:F{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    for p in user_pages:
        exists = p.get("exists", False)
        sc     = p.get("status_code", 0)
        if exists:
            bg_e, ft_e, badge = O_BG, O_FT, "EXISTUJE ⚠"
        elif sc == 404 or sc == 0:
            bg_e, ft_e, badge = G_BG, G_FT, "Neexistuje"
        else:
            bg_e, ft_e, badge = GR_BG, GR_FT, f"HTTP {sc}"

        _dc(ws, row, 1, p["path"], bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        _dc(ws, row, 2, _as_https(p["url"]))
        ws.merge_cells(f"E{row}:F{row}")
        _dc(ws, row, 5, sc if sc else "–", align="center",
            bg=O_BG if exists else GR_BG)
        ws.merge_cells(f"G{row}:G{row}")
        _badge(ws, row, 7, badge, bg_e, ft_e)
        ws.row_dimensions[row].height = 18; row += 1

    return row


# ── Hlavní funkce ─────────────────────────────────────────────────────────────

def write_report(results: list, output_path: Path, start_url: str,
                 score: int = -1, source_label: str = "",
                 domain_info: dict | None = None) -> None:
    """Generuje Excel report."""
    if domain_info is None:
        domain_info = {}

    stats          = compute_stats(results)
    robots_issues  = domain_info.get("robots_issues", [])
    robots_skipped = domain_info.get("robots_skipped", False)
    user_pages     = domain_info.get("user_pages", [])

    is_local_audit = is_local_url(start_url)

    # Detekce jestli se nějaká URL převedla z http na https.
    # Pro lokální URL (localhost…) http:// nepřevádíme — http_converted by
    # pak ukazoval falešné upozornění o konverzi která se ve skutečnosti nestala.
    def _is_converted_http(u: str) -> bool:
        return isinstance(u, str) and u.startswith("http://") and not is_local_url(u)

    http_converted = (
        any(_is_converted_http(r.get("url", "")) for r in results)
        or _is_converted_http(start_url)
        or any(_is_converted_http(p.get("url", "")) for p in user_pages)
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("📋 Report")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFG", [80.69, 12.51, 8.76, 8.0, 30, 14, 40]):
        ws.column_dimensions[col].width = w

    row = _write_header(ws, start_url, source_label)
    row = _write_summary(ws, row, score, stats, http_converted=http_converted)
    row = _write_homepage_meta(ws, row, results)
    row = _write_w3c_section(ws, row, results, is_local_audit=is_local_audit)
    row = _write_structure_section(ws, row, results)
    row = _write_failed_pages(ws, row, results)
    row = _write_robots_section(ws, row, robots_issues, robots_skipped)
    row = _write_user_pages(ws, row, user_pages, is_local_audit=is_local_audit)

    ws.freeze_panes = "A4"
    wb.save(output_path)